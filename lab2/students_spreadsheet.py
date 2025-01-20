import openpyxl

# Функция для чтения Excel-файла
def load_attendance(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Считываем названия дисциплин
    disciplines = [sheet.cell(row=1, column=col).value for col in range(5, sheet.max_column + 1)]

    # Считываем даты
    dates = [sheet.cell(row=2, column=col).value for col in range(5, sheet.max_column + 1)]

    # Считываем студентов
    students = []
    for row in range(3, 23):  # строки 3-22 (20 студентов)
        student_id = sheet.cell(row=row, column=1).value
        last_name = sheet.cell(row=row, column=2).value
        first_name = sheet.cell(row=row, column=3).value
        middle_name = sheet.cell(row=row, column=4).value

        # Посещаемость
        attendance = {}
        for col in range(5, sheet.max_column + 1):
            discipline = disciplines[col - 5]
            date = dates[col - 5]
            status = sheet.cell(row=row, column=col).value
            attendance[(discipline, date)] = status if status else "был"

        students.append({
            "id": student_id,
            "last_name": last_name,
            "first_name": first_name,
            "middle_name": middle_name,
            "attendance": attendance
        })

    return students

# Функция для вывода данных о студенте
def print_student_info(student):
    print(f"ID: {student['id']}")
    print(f"Фамилия: {student['last_name']}")
    print(f"Имя: {student['first_name']}")
    print(f"Отчество: {student['middle_name']}")
    print("Посещаемость:")
    for (discipline, date), status in student["attendance"].items():
        print(f"  {discipline} ({date}): {status}")

# Функция для поиска студента по ID
def find_student_by_id(students, student_id):
    for student in students:
        if student["id"] == student_id:
            return student
    return None

# Функция для изменения статуса посещаемости
def update_attendance(student, discipline, date, new_status):
    key = (discipline, date)
    if key in student["attendance"]:
        student["attendance"][key] = new_status
        print(f"Посещаемость обновлена: {discipline} ({date}) -> {new_status}")
    else:
        print(f"Данные для дисциплины '{discipline}' и даты '{date}' не найдены.")


if __name__ == "__main__":
    # Путь к файлу
    file_path = "students.xlsx"
    
    # Загрузка данных
    students = load_attendance(file_path)
    print("Данные загружены.")

    # Работа с данными
    while True:
        print("\nМеню:")
        print("1) Показать данные студента по ID")
        print("2) Изменить статус посещаемости")
        print("3) Выйти")
        choice = input("Выберите действие: ")

        if choice == "1":
            student_id = int(input("Введите ID студента: "))
            student = find_student_by_id(students, student_id)
            if student:
                print_student_info(student)
            else:
                print("Студент не найден.")
        
        elif choice == "2":
            student_id = int(input("Введите ID студента: "))
            student = find_student_by_id(students, student_id)
            if student:
                discipline = input("Введите название дисциплины: ")
                date = input("Введите дату (например, 01.янв): ")
                new_status = input("Введите новый статус ('н', 'н/у', 'б' или пусто): ")
                update_attendance(student, discipline, date, new_status)
            else:
                print("Студент не найден.")
        
        elif choice == "3":
            print("Выход из программы.")
            break
        
        else:
            print("Некорректный выбор.")
